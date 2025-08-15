import re
from unidecode import unidecode
import pandas as pd
import openpyxl
import unicodedata
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.http import require_POST
from unidecode import unidecode
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .models import Usuario, Campus, GrupoTrabalho, Cidade, Estado, Edital, Avaliacao
from .forms import (
    UsuarioForm, CidadeForm, EstadoForm, CampusForm,
    GrupoTrabalhoForm, EditalForm, GrupoAtendimentoForm
)

# ==================== USUÁRIOS ====================


@login_required
def remover_acentos(texto):
    if texto is None:
        return ''
    nfkd_form = unicodedata.normalize('NFKD', texto)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])


@login_required
def lista_usuarios(request):
    usuarios = Usuario.objects.all()

    termo_busca_nome = request.GET.get('nome')
    termo_busca_cpf = request.GET.get('cpf')
    termo_busca_campus = request.GET.getlist('campus')
    termo_busca_grupos = request.GET.getlist('grupos')
    termo_busca_ativo = request.GET.get('ativo')

    if termo_busca_cpf:
        cpf_limpo = re.sub(r'\D', '', termo_busca_cpf)
        usuarios = usuarios.filter(cpf__icontains=cpf_limpo)

    if termo_busca_campus and termo_busca_campus != ['']:
        usuarios = usuarios.filter(campus__id__in=termo_busca_campus)

    if termo_busca_grupos and termo_busca_grupos != ['']:
        usuarios = usuarios.filter(
            grupos__id__in=termo_busca_grupos).distinct()

    if termo_busca_ativo is not None:
        if termo_busca_ativo.lower() == 'sim':
            usuarios = usuarios.filter(ativo=True)
        elif termo_busca_ativo.lower() == 'nao':
            usuarios = usuarios.filter(ativo=False)

    if termo_busca_nome:
        termo_busca_normalizado = remover_acentos(termo_busca_nome).lower()
        usuarios = [
            u for u in usuarios if termo_busca_normalizado in unidecode(u.nome).lower()]
    else:
        usuarios = usuarios.order_by('nome')

    # Busca campi e grupos para os filtros no template
    campi = Campus.objects.all().order_by('nome')
    grupos = GrupoTrabalho.objects.all().order_by('nome')

    context = {
        'usuarios': usuarios,
        'campi': campi,
        'grupos': grupos,
        'termo_busca_nome': termo_busca_nome,
        'termo_busca_cpf': termo_busca_cpf,
        'termo_busca_campus': termo_busca_campus,
        'termo_busca_grupos': termo_busca_grupos,
        'ativo_filtro': termo_busca_ativo,  # <-- Adicione isto!
    }

    return render(request, 'usuarios/lista_usuarios.html', context)


@login_required
def adicionar_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.delete()
    return redirect('lista_usuarios')

# ==================== CIDADES ====================


@login_required
def lista_cidades(request):
    cidades = Cidade.objects.all()

    # Adicionar o filtro
    termo_busca_nome = request.GET.get('nome')
    termo_busca_estado = request.GET.get('estado')

    if termo_busca_nome:
        cidades = cidades.filter(nome__icontains=termo_busca_nome)

    if termo_busca_estado:
        # A busca por estado precisa do ID ou do nome. Vamos usar o ID.
        cidades = cidades.filter(estado__id=termo_busca_estado)

    # Buscar todos os estados para popular o seletor no template
    estados = Estado.objects.all().order_by('nome')

    # Passar a lista de estados para o template
    context = {
        'cidades': cidades,
        'estados': estados,
        'termo_busca_nome': termo_busca_nome,
        'termo_busca_estado': termo_busca_estado,
    }

    return render(request, 'lista_cidades.html', context)


@login_required
def adicionar_cidade(request):
    if request.method == 'POST':
        form = CidadeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_cidades')
    else:
        form = CidadeForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_cidade(request, cidade_id):
    cidade = get_object_or_404(Cidade, id=cidade_id)
    if request.method == 'POST':
        form = CidadeForm(request.POST, instance=cidade)
        if form.is_valid():
            form.save()
            return redirect('lista_cidades')
    else:
        form = CidadeForm(instance=cidade)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_cidade(request, cidade_id):
    cidade = get_object_or_404(Cidade, id=cidade_id)
    cidade.delete()
    return redirect('lista_cidades')

# ==================== ESTADOS ====================


@login_required
def lista_estados(request):
    estados = Estado.objects.all()
    return render(request, 'lista_estados.html', {'estados': estados})


@login_required
def adicionar_estado(request):
    if request.method == 'POST':
        form = EstadoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_estados')
    else:
        form = EstadoForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_estado(request, estado_id):
    estado = get_object_or_404(Estado, id=estado_id)
    if request.method == 'POST':
        form = EstadoForm(request.POST, instance=estado)
        if form.is_valid():
            form.save()
            return redirect('lista_estados')
    else:
        form = EstadoForm(instance=estado)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_estado(request, estado_id):
    estado = get_object_or_404(Estado, id=estado_id)
    estado.delete()
    return redirect('lista_estados')

# ==================== CAMPUS ====================


@login_required
def lista_campus(request):
    campus = Campus.objects.all()
    return render(request, 'lista_campus.html', {'campus': campus})


@login_required
def adicionar_campus(request):
    if request.method == 'POST':
        form = CampusForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_campus')
    else:
        form = CampusForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_campus(request, campus_id):
    campus = get_object_or_404(Campus, id=campus_id)
    if request.method == 'POST':
        form = CampusForm(request.POST, instance=campus)
        if form.is_valid():
            form.save()
            return redirect('lista_campus')
    else:
        form = CampusForm(instance=campus)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_campus(request, campus_id):
    campus = get_object_or_404(Campus, id=campus_id)
    campus.delete()
    return redirect('lista_campus')

# ==================== GRUPOS DE TRABALHO ====================


@login_required
def lista_grupos(request):
    grupos = GrupoTrabalho.objects.all()
    return render(request, 'lista_grupos.html', {'grupos': grupos})


@login_required
def adicionar_grupo(request):
    if request.method == 'POST':
        form = GrupoTrabalhoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_grupos')
    else:
        form = GrupoTrabalhoForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_grupo(request, grupo_id):
    grupo = get_object_or_404(GrupoTrabalho, id=grupo_id)
    if request.method == 'POST':
        form = GrupoTrabalhoForm(request.POST, instance=grupo)
        if form.is_valid():
            form.save()
            return redirect('lista_grupos')
    else:
        form = GrupoTrabalhoForm(instance=grupo)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_grupo(request, grupo_id):
    grupo = get_object_or_404(GrupoTrabalho, id=grupo_id)
    grupo.delete()
    return redirect('lista_grupos')


@login_required
def dashboard(request):
    return render(request, 'usuarios/dashboard.html')

# ==================== EXPORTAÇÃO / IMPORTAÇÃO ====================


@login_required
def exportar_usuarios(request):
    usuarios = Usuario.objects.select_related('campus__cidade__estado').all()
    dados = []

    for u in usuarios:
        dados.append({
            'Nome': u.nome,
            'CPF': u.cpf,
            'Matrícula': u.matricula,
            'Campus': u.campus.nome,
            'Email': u.email,
            'Endereço': u.endereco,
            'Banco': u.banco,
            'Agencia': u.agencia,
            'Conta': u.conta,
            'Chave Pix': u.chave_pix,
            'Cidade': u.campus.cidade.nome,
            'Estado': u.campus.cidade.estado.nome,
            'Bairro': u.bairro,
            'Cep': u.cep,
            'Telefone': u.telefone,
            'Ativo': u.ativo,
            'Grupos': ", ".join([g.nome for g in u.grupos.all()])
        })

    df = pd.DataFrame(dados)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    df.to_excel(response, index=False)
    return response


@login_required
def exportar_grupos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grupos"

    # Cabeçalhos
    ws.append(['Nome do Grupo'])

    # Dados
    for grupo in GrupoTrabalho.objects.all():
        ws.append([grupo.nome])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=grupos.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_cidades(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cidades"

    # Cabeçalhos
    ws.append(['Nome da Cidade', 'Estado'])

    # Dados
    for cidade in Cidade.objects.select_related('estado').all():
        ws.append([cidade.nome, cidade.estado.nome])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cidades.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_estados(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estados"

    # Cabeçalhos
    ws.append(['Nome do Estado'])

    # Dados
    for estado in Estado.objects.all():
        ws.append([estado.nome])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=estados.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_campus(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campus"

    # Cabeçalhos
    ws.append(['Nome do Campus', 'Sigla do Campus'])

    # Dados
    for campus in Campus.objects.all():
        ws.append([campus.nome, campus.sigla])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=campus.xlsx'
    wb.save(response)
    return response


@login_required
def importar_excel(request):
    if request.method == 'POST' and request.FILES['arquivo']:
        df = pd.read_excel(request.FILES['arquivo'])

        for _, row in df.iterrows():
            estado, _ = Estado.objects.get_or_create(nome=row['Estado'])
            cidade, _ = Cidade.objects.get_or_create(
                nome=row['Cidade'], estado=estado)
            campus, _ = Campus.objects.get_or_create(
                nome=row['Campus'], cidade=cidade)

            usuario = Usuario.objects.create(
                nome=row['Nome'],
                cpf=row['CPF'],
                matricula=row['Matrícula'],
                chave_pix=row['Chave Pix'],
                conta=row['Conta'],
                agencia=row['Agencia'],
                endereco=row['Endereço'],
                bairro=row['Bairro'],
                cep=row['Cep'],
                telefone=row['Telefone'],
                campus=campus
            )

            grupos = row.get('Grupos', '').split(',')
            for nome_grupo in grupos:
                nome_grupo = nome_grupo.strip()
                if nome_grupo:
                    grupo, _ = GrupoTrabalho.objects.get_or_create(
                        nome=nome_grupo)
                    usuario.grupos.add(grupo)

        return redirect('lista_usuarios')

    return render(request, 'cadastro/lista.html', {'usuarios': Usuario.objects.all()})


# ==================== EDITAIS ====================
@login_required
def lista_edital(request):
    editais = Edital.objects.all()

    termo_nome = request.GET.get('nome')
    termo_numero = request.GET.get('numero')
    termo_campus = request.GET.get('campus')
    termo_ativo = request.GET.get('ativo')

    if termo_ativo is not None:
        if termo_ativo.lower() == 'sim':
            editais = editais.filter(ativo=True)
        elif termo_ativo.lower() == 'nao':
            editais = editais.filter(ativo=False)

    if termo_nome:
        editais = editais.filter(nome__icontains=termo_nome)

    if termo_numero:
        editais = editais.filter(numero__icontains=termo_numero)

    if termo_campus:
        editais = editais.filter(campus__nome__icontains=termo_campus)

    context = {
        'editais': editais,
    }
    return render(request, 'lista_editais.html', context)


@login_required
def adicionar_edital(request):
    if request.method == 'POST':
        form = EditalForm(request.POST)
        if form.is_valid():
            form.save()
            # redireciona para a lista de editais
            return redirect('lista_editais')
    else:
        form = EditalForm()
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def editar_edital(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)
    if request.method == 'POST':
        form = EditalForm(request.POST, instance=edital)
        if form.is_valid():
            form.save()
            return redirect('lista_editais')
    else:
        form = EditalForm(instance=edital)
    return render(request, 'usuarios/formulario.html', {'form': form})


@login_required
def excluir_edital(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)
    edital.delete()
    return redirect('lista_editais')


@login_required
def exportar_editais(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Editais"

    # Cabeçalhos
    ws.append([
        'Nome do Edital', 'Numero', 'Campus', 'Avaliadores Historico',
        'Avaliadores Renda', 'Ativo', 'ID'
    ])
    # Dados
    for edital in Edital.objects.all():
        campus_nomes = ", ".join([c.nome for c in edital.campus.all()])
        avaliadores_historico = ", ".join(
            [a.nome for a in edital.avaliadores_historico.all()])
        avaliadores_renda = ", ".join(
            [a.nome for a in edital.avaliadores_renda.all()])

        ws.append([
            edital.nome,
            edital.numero,
            campus_nomes,
            avaliadores_historico,
            avaliadores_renda,
            edital.ativo,
            edital.id,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=editais.xlsx'
    wb.save(response)
    return response


@login_required
def visualizar_edital(request, edital_id):
    edital = get_object_or_404(Edital, pk=edital_id)

    if request.method == 'POST':
        form = EditalForm(request.POST, instance=edital)
        if form.is_valid():
            form.save()
            return redirect('visualizar_edital', edital_id=edital.pk)
    else:
        form = EditalForm(instance=edital)

    return render(request, 'visualizar_edital.html', {'form': form, 'edital': edital})


@login_required
def importar_equipe_edital(request, id):
    edital = get_object_or_404(Edital, id=id)

    grupos = edital.grupos.all()

    usuarios_importados = 0
    for grupo in grupos:
        usuarios = grupo.usuarios.all()
        for usuario in usuarios:
            if not edital.equipe.filter(id=usuario.id).exists():
                edital.equipe.add(usuario)
                usuarios_importados += 1

    messages.success(
        request, f'{usuarios_importados} usuários importados para a equipe do edital.')
    return redirect('visualizar_edital', edital_id=edital.id)


@login_required
def adicionar_grupo_atendimento(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)

    if request.method == 'POST':
        form = GrupoAtendimentoForm(request.POST)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.edital = edital
            grupo.save()
            return redirect('visualizar_edital', edital_id=edital.id)
    else:
        form = GrupoAtendimentoForm()

    return render(request, 'adicionar_grupo_atendimento.html', {'form': form, 'edital': edital})


@login_required
def selecionar_grupo_trabalho(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)
    grupos = grupos = edital.grupos.all()
    return render(request, 'selecionar_grupo_trabalho.html', {'edital': edital, 'grupos': grupos})


@login_required
def visualizar_edital_grupo(request, edital_id, grupo_id):
    edital = get_object_or_404(Edital, id=edital_id)
    grupo = get_object_or_404(GrupoTrabalho, id=grupo_id, editais=edital)
    return render(request, 'editais/visualizar_edital_grupo.html', {'edital': edital, 'grupo': grupo})


@login_required
def visualizar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    return render(request, 'visualizar_usuario.html', {'usuario': usuario})


@login_required
def alternar_ativo_edital(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)
    edital.ativo = not edital.ativo
    edital.save()

    if edital.ativo:
        messages.success(request, "Edital ativado com sucesso.")
    else:
        messages.warning(request, "Edital inativado com sucesso.")

    return redirect('lista_editais')


@require_POST
def atualizar_status_editais(request):
    ids_ativos = request.POST.getlist('ativos')
    todos_editais = Edital.objects.all()

    for edital in todos_editais:
        edital.ativo = str(edital.id) in ids_ativos
        edital.save()

    messages.success(request, "Status dos editais atualizado com sucesso.")
    return redirect('lista_editais')


@login_required
def importar_estados(request):
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        arquivo_excel = request.FILES['arquivo_excel']

        try:
            df = pd.read_excel(arquivo_excel)  # Lê a planilha
            for _, row in df.iterrows():
                nome_estado = row['Nome'].strip()
                if not Estado.objects.filter(nome=nome_estado).exists():
                    Estado.objects.create(nome=nome_estado)

            messages.success(request, 'Importação realizada com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao importar: {str(e)}')

        return redirect('lista_estados')

    return render(request, 'importar_estados.html')


@login_required
def importar_cidades(request):
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        arquivo_excel = request.FILES['arquivo_excel']

        try:
            df = pd.read_excel(arquivo_excel)  # Lê o Excel

            for _, row in df.iterrows():
                nome_cidade = str(row['Nome']).strip()
                nome_estado = str(row['Estado']).strip()

                # Busca o estado correspondente
                estado = Estado.objects.filter(
                    nome__iexact=nome_estado).first()
                if not estado:
                    messages.warning(
                        request, f"Estado '{nome_estado}' não encontrado. Cidade '{nome_cidade}' ignorada.")
                    continue

                # Cria a cidade se ainda não existir no mesmo estado
                if not Cidade.objects.filter(nome=nome_cidade, estado=estado).exists():
                    Cidade.objects.create(nome=nome_cidade, estado=estado)

            messages.success(request, 'Cidades importadas com sucesso!')

        except Exception as e:
            messages.error(request, f'Erro ao importar: {str(e)}')

        return redirect('lista_cidades')

    return render(request, 'importar_cidades.html')


@login_required
def importar_campus(request):
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        arquivo_excel = request.FILES['arquivo_excel']

        try:
            df = pd.read_excel(arquivo_excel)  # Lê a planilha

            for _, row in df.iterrows():
                nome_campus = str(row['Nome']).strip()
                sigla_campus = str(row['Sigla']).strip()
                cidade_nome = str(row['Cidade']).strip()

                cidade_obj = Cidade.objects.filter(
                    nome__iexact=cidade_nome).first()

                if not cidade_obj:
                    messages.warning(
                        request, f"Cidade '{cidade_nome}' não encontrada para o campus '{nome_campus}'. Linha ignorada.")
                    continue  # pula para próxima linha

                campus = Campus.objects.filter(
                    sigla__iexact=sigla_campus).first()

                if campus:
                    # Atualiza o campus se cidade for diferente
                    if campus.cidade != cidade_obj:
                        campus.cidade = cidade_obj
                        campus.nome = nome_campus  # opcional, caso queira atualizar nome também
                        campus.save()
                else:
                    # Cria campus com cidade associada
                    Campus.objects.create(
                        nome=nome_campus,
                        sigla=sigla_campus,
                        cidade=cidade_obj
                    )

            messages.success(request, 'Campi importados com sucesso!')

        except Exception as e:
            messages.error(request, f'Erro ao importar: {str(e)}')

        return redirect('lista_campus')

    return render(request, 'importar_campus.html')


def normalizar(texto):
    """Remove acentos, converte para minúsculas e tira espaços extras."""
    if pd.isna(texto):
        return ''
    texto = str(texto).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(c))


@login_required
def importar_usuarios(request):
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        arquivo_excel = request.FILES['arquivo_excel']

        try:
            df = pd.read_excel(arquivo_excel)

            # Normaliza o nome das colunas
            df.columns = [normalizar(col) for col in df.columns]

            for _, row in df.iterrows():
                nome = str(row.get('nome', '')).strip()
                cpf = str(row.get('cpf', '')).strip()
                matricula = str(row.get('matricula', '')).strip()
                email = str(row.get('email', '')).strip()
                endereco = str(row.get('endereco', '')).strip()
                chave_pix = str(row.get('chave pix', '')).strip()

                # Campus
                campus_nome = str(row.get('campus', '')).strip()
                campus, _ = Campus.objects.get_or_create(
                    nome__iexact=campus_nome, defaults={'nome': campus_nome})

                # Grupos (aceita vírgula ou ;)
                grupos_nomes = str(row.get('grupos', '')).replace(
                    ';', ',').split(',')
                grupos = []
                for gnome in grupos_nomes:
                    gnome = gnome.strip()
                    if not gnome:
                        continue
                    grupo, _ = GrupoTrabalho.objects.get_or_create(
                        nome__iexact=gnome, defaults={'nome': gnome})
                    grupos.append(grupo)

                banco = str(row.get('banco', '')).strip()
                agencia = str(row.get('agencia', '')).strip()
                conta = str(row.get('conta', '')).strip()
                bairro = str(row.get('bairro', '')).strip()

                # Estado e Cidade
                estado_nome = str(row.get('estado', '')).strip()
                estado, _ = Estado.objects.get_or_create(
                    nome__iexact=estado_nome, defaults={'nome': estado_nome})

                cidade_nome = str(row.get('cidade', '')).strip()
                cidade, _ = Cidade.objects.get_or_create(nome__iexact=cidade_nome, estado=estado, defaults={
                                                         'nome': cidade_nome, 'estado': estado})

                cep = str(row.get('cep', '')).strip()
                telefone = str(row.get('telefone', '')).strip()
                ativo = True if str(row.get('ativo', '')).lower() in [
                    'sim', 'yes', 'true', '1'] else False
                recebeu_email = True if str(row.get('recebeu email', '')).lower() in [
                    'sim', 'yes', 'true', '1'] else False

                # Observações (aceita singular/plural)
                observacoes = ''
                if 'observacoes' in df.columns:
                    observacoes = str(row.get('observacoes', '')).strip()
                elif 'observacao' in df.columns:
                    observacoes = str(row.get('observacao', '')).strip()

                # Cria ou atualiza usuário
                usuario, _ = Usuario.objects.update_or_create(
                    cpf=cpf,
                    defaults={
                        'nome': nome,
                        'matricula': matricula,
                        'email': email,
                        'endereco': endereco,
                        'chave_pix': chave_pix,
                        'campus': campus,
                        'banco': banco,
                        'agencia': agencia,
                        'conta': conta,
                        'bairro': bairro,
                        'cidade': cidade,
                        'estado': estado,
                        'cep': cep,
                        'telefone': telefone,
                        'ativo': ativo,
                        'recebeu_email': recebeu_email,
                        'observacoes': observacoes,
                    }
                )
                usuario.grupos.set(grupos)

            messages.success(
                request, 'Importação de usuários realizada com sucesso!')

        except Exception as e:
            messages.error(request, f'Erro ao importar: {str(e)}')

        return redirect('lista_usuarios')

    return render(request, 'importar_usuarios.html')


@login_required
def lista_avaliadores(request, edital_id):
    VALOR_POR_AVALIACAO = 5.20

    avaliadores = Usuario.objects.filter(
        avaliacao__edital_id=edital_id
    ).distinct()

    dados_tabela = []

    for avaliador in avaliadores:
        qtd_avaliacoes = Avaliacao.objects.filter(
            avaliador=avaliador,
            edital_id=edital_id,
            tipo="analise"
        ).count()

        qtd_recursos = Avaliacao.objects.filter(
            avaliador=avaliador,
            edital_id=edital_id,
            tipo="recurso"
        ).count()

        total = qtd_avaliacoes + qtd_recursos
        total_receber = total * VALOR_POR_AVALIACAO

        dados_tabela.append({
            "nome": avaliadores.get_full_name(),  # ou username, dependendo do seu model
            "avaliacoes": qtd_avaliacoes,
            "recursos": qtd_recursos,
            "total": total,
            "valor_unitario": VALOR_POR_AVALIACAO,
            "total_receber": total_receber,
        })

    return render(request, "cadastro/lista_avaliadores.html", {
        "dados_tabela": dados_tabela,
        "edital_id": edital_id
    })


@login_required
def detalhes_edital(request, edital_id):
    edital = get_object_or_404(Edital, id=edital_id)

    VALOR_POR_AVALIACAO = 5.20

    avaliadores = Usuario.objects.filter(
        avaliacao__edital_id=edital_id
    ).distinct()

    dados_tabela = []

    for avaliador in avaliadores:
        qtd_avaliacoes = Avaliacao.objects.filter(
            avaliador=avaliador,
            edital_id=edital_id,
            tipo="analise"
        ).count()

        qtd_recursos = Avaliacao.objects.filter(
            avaliador=avaliador,
            edital_id=edital_id,
            tipo="recurso"
        ).count()

        total = qtd_avaliacoes + qtd_recursos
        total_receber = total * VALOR_POR_AVALIACAO

        dados_tabela.append({
            "nome": avaliador.get_full_name(),  # ou username, ou nome
            "avaliacoes": qtd_avaliacoes,
            "recursos": qtd_recursos,
            "total": total,
            "valor_unitario": VALOR_POR_AVALIACAO,
            "total_receber": total_receber,
        })

    return render(request, 'cadastro/selecionar_grupo_trabalho.html', {
        'edital': edital,
        'dados_tabela': dados_tabela
    })
